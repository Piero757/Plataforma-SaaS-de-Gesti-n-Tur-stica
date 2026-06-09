from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, Count, F, Q
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import (
    Venta, Producto, Inventario, RegistroServicio, Cliente, Proveedor, 
    Compra, FacturaElectronica, DetalleVenta, ConfiguracionEmpresa,
    Habitacion, Mesa, Reserva, PedidoHabitacion, DetalleCompra, CuotaCompra,
    CategoriaPersonalizada, PerfilUsuario, EstadoComanda
)
from django.contrib.auth import login
from .forms import (
    ClienteForm, UsuarioCreateForm, RegistrationForm, ProveedorForm, ProductoForm, 
    CompraForm, VentaForm, ServicioForm, ConfiguracionEmpresaForm,
    HabitacionForm, MesaForm, ReservaForm, PedidoHabitacionForm, DetalleCompraForm, CuotaCompraForm,
    UserProfileForm
)
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
import openpyxl
import csv
from datetime import datetime
from .utils import render_to_pdf
from django.contrib.auth.models import User
from django.db.models.functions import TruncMonth
import json
from decimal import Decimal


# --- DASHBOARD ---

@login_required
def dashboard_view(request):
    # Redirección automática según el rol del usuario
    try:
        rol = request.user.perfil.rol
    except PerfilUsuario.DoesNotExist:
        rol = 'ADMIN'

    if rol == 'COCINA':
        return redirect('cocina')
    elif rol == 'BARRA':
        return redirect('barra')
    elif rol == 'LIMPIEZA':
        return redirect('limpieza')
    elif rol in ['MOZO', 'JEFE_MOZO']:
        return redirect('restaurante')

    today = timezone.now().date()
    month_start = today.replace(day=1)

    modulo = request.session.get('modulo', 'HOTEL')
    
    if modulo == 'RESTAURANTE':
        return redirect('restaurante')
    
    total_usuarios = User.objects.count()
    ventas_hoy = Venta.objects.filter(fecha__date=today, modulo=modulo).aggregate(Sum('total'))['total__sum'] or 0
    productos_inventario = Producto.objects.filter(modulo=modulo).count()
    servicios_realizados = RegistroServicio.objects.filter(modulo=modulo).count()
    ingresos_mes = Venta.objects.filter(fecha__date__gte=month_start, modulo=modulo).aggregate(Sum('total'))['total__sum'] or 0

    context = {
        'modulo_actual': modulo,
        'total_usuarios': total_usuarios,
        'ventas_hoy': ventas_hoy,
        'productos_inventario': productos_inventario,
        'servicios_realizados': servicios_realizados,
        'ingresos_mes': ingresos_mes,
        'ventas_recientes': Venta.objects.filter(modulo=modulo).order_by('-fecha')[:5],
    }
    return render(request, 'gestion/dashboard.html', context)

@login_required
def restaurante_view(request):
    today = timezone.now().date()
    # Productos de restaurante
    productos = Producto.objects.filter(categoria='RESTAURANTE', activo=True)
    
    # Ventas de restaurante (aquellas que incluyen al menos un producto de restaurante)
    ventas_restaurante = Venta.objects.filter(
        detalles__producto__categoria='RESTAURANTE'
    ).distinct().order_by('-fecha')
    
    # Métricas hoy
    ventas_hoy = Venta.objects.filter(
        fecha__date=today,
        detalles__producto__categoria='RESTAURANTE'
    ).distinct().aggregate(Sum('total'))['total__sum'] or 0
    
    mesas = Mesa.objects.all().order_by('numero')
    
    context = {
        'productos': productos,
        'ventas_recientes': ventas_restaurante[:5],
        'ventas_hoy': ventas_hoy,
        'cantidad_productos': productos.count(),
        'mesas': mesas,
        'titulo': 'Panel de Restaurante'
    }
    return render(request, 'gestion/restaurante_dashboard.html', context)

@login_required
def cambiar_modulo(request):
    modulo = request.GET.get('modulo', 'HOTEL')
    if modulo in ['HOTEL', 'RESTAURANTE']:
        request.session['modulo'] = modulo
        messages.success(request, f"Cambiado al módulo: {modulo.capitalize()}")
    return redirect('dashboard')

# --- USUARIOS ---

@login_required
def usuarios_list(request):
    usuarios = User.objects.all()
    return render(request, 'gestion/usuarios_list.html', {'usuarios': usuarios})

@login_required
def usuario_create(request):
    if request.method == 'POST':
        form = UsuarioCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente.")
            return redirect('usuarios_list')
    else:
        form = UsuarioCreateForm()
    return render(request, 'gestion/usuario_form.html', {'form': form, 'titulo': 'Crear Usuario'})

@login_required
def usuario_update(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UsuarioCreateForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect('usuarios_list')
    else:
        form = UsuarioCreateForm(instance=usuario)
    return render(request, 'gestion/usuario_form.html', {'form': form, 'titulo': 'Editar Usuario'})

@login_required
def usuario_toggle_status(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if usuario == request.user:
        messages.error(request, "No puedes desactivarte a ti mismo.")
    else:
        usuario.is_active = not usuario.is_active
        usuario.save()
        status = "activado" if usuario.is_active else "desactivado"
        messages.info(request, f"Usuario {status} correctamente.")
    return redirect('usuarios_list')

@login_required
def perfil_view(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return redirect('perfil')
    else:
        form = UserProfileForm(instance=request.user)
    return render(request, 'gestion/perfil.html', {'form': form, 'titulo': 'Mi Perfil'})

class CustomPasswordChangeView(PasswordChangeView):
    template_name = 'gestion/perfil_password.html'
    success_url = reverse_lazy('perfil')
    
    def form_valid(self, form):
        messages.success(self.request, "Contraseña actualizada correctamente.")
        return super().form_valid(form)

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    empresa_config = ConfiguracionEmpresa.objects.first()
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f"¡Bienvenido {user.username}! Te has registrado correctamente.")
            return redirect('dashboard')
    else:
        form = RegistrationForm()
    
    return render(request, 'registration/register.html', {
        'form': form,
        'empresa_config': empresa_config,
        'titulo': 'Registrarse'
    })

@login_required
def usuario_toggle_admin(request, pk):
    if not request.user.is_superuser:
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('usuarios_list')
        
    usuario = get_object_or_404(User, pk=pk)
    if usuario == request.user:
        messages.error(request, "No puedes cambiar tu propio nivel de acceso.")
    else:
        # Toggle both is_staff and is_superuser for full admin access
        is_admin = not (usuario.is_staff and usuario.is_superuser)
        usuario.is_staff = is_admin
        usuario.is_superuser = is_admin
        usuario.save()
        status = "promovido a Administrador" if is_admin else "cambiado a Operador"
        messages.success(request, f"Usuario {usuario.username} {status}.")
        
    return redirect('usuarios_list')

from django.db.models import Q

# --- CLIENTES ---

@login_required
def clientes_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    query = request.GET.get('q')
    tipo = request.GET.get('tipo')
    
    clientes = Cliente.objects.filter(activo=True, modulo=modulo).annotate(
        num_hospedajes=Count('reserva', distinct=True)
    )
    
    if query:
        clientes = clientes.filter(
            Q(nombre_razon_social__icontains=query) | 
            Q(numero_documento__icontains=query)
        )
    
    if tipo and tipo != 'Todos':
        clientes = clientes.filter(tipo_documento=tipo)
        
    context = {
        'clientes': clientes,
        'query': query,
        'tipo_seleccionado': tipo,
        'tipos_documento': Cliente.TIPO_DOCUMENTO
    }
    return render(request, 'gestion/clientes_list.html', context)

@login_required
def clientes_historial(request):
    query = request.GET.get('q')
    tipo = request.GET.get('tipo')
    
    clientes = Cliente.objects.filter(activo=False).annotate(
        num_hospedajes=Count('reserva', distinct=True)
    )
    
    if query:
        clientes = clientes.filter(
            Q(nombre_razon_social__icontains=query) | 
            Q(numero_documento__icontains=query)
        )
    
    if tipo and tipo != 'Todos':
        clientes = clientes.filter(tipo_documento=tipo)

    return render(request, 'gestion/clientes_list.html', {
        'clientes': clientes, 
        'es_historial': True,
        'titulo': 'Historial de Clientes (Eliminados)',
        'query': query,
        'tipo_seleccionado': tipo,
        'tipos_documento': Cliente.TIPO_DOCUMENTO
    })

@login_required
def cliente_create(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.modulo = request.session.get('modulo', 'HOTEL')
            cliente.save()
            messages.success(request, "Cliente registrado correctamente.")
            return redirect('clientes_list')
    else:
        form = ClienteForm()
    return render(request, 'gestion/cliente_form.html', {'form': form, 'titulo': 'Nuevo Cliente'})

@login_required
def cliente_update(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente actualizado.")
            return redirect('clientes_list')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'gestion/cliente_form.html', {'form': form, 'titulo': 'Editar Cliente'})

@login_required
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = False
    cliente.save()
    messages.success(request, "Cliente movido al historial.")
    return redirect('clientes_list')

@login_required
def cliente_restore(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = True
    cliente.save()
    messages.success(request, "Cliente restaurado correctamente.")
    return redirect('clientes_list')

# --- PROVEEDORES ---

@login_required
def proveedores_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    proveedores = Proveedor.objects.filter(activo=True, modulo=modulo)
    return render(request, 'gestion/proveedores_list.html', {'proveedores': proveedores})

@login_required
def proveedores_historial(request):
    proveedores = Proveedor.objects.filter(activo=False)
    return render(request, 'gestion/proveedores_list.html', {
        'proveedores': proveedores,
        'es_historial': True,
        'titulo': 'Historial de Proveedores'
    })

@login_required
def proveedor_create(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            prov = form.save(commit=False)
            prov.modulo = request.session.get('modulo', 'HOTEL')
            prov.save()
            messages.success(request, "Proveedor registrado.")
            return redirect('proveedores_list')
    else:
        form = ProveedorForm()
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Nuevo Proveedor'})

@login_required
def proveedor_update(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor actualizado.")
            return redirect('proveedores_list')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Editar Proveedor'})

@login_required
def proveedor_delete(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    proveedor.activo = False
    proveedor.save()
    messages.success(request, "Proveedor movido al historial.")
    return redirect('proveedores_list')

@login_required
def proveedor_restore(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    proveedor.activo = True
    proveedor.save()
    messages.success(request, "Proveedor restaurado.")
    return redirect('proveedores_list')

# --- PRODUCTOS / INVENTARIO ---

@login_required
def inventario_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    inventario = Inventario.objects.filter(producto__activo=True, producto__modulo=modulo).order_by('producto__subcategoria', 'producto__nombre')
    
    # Stock critico: actual <= minimo
    stock_critico_count = inventario.filter(stock_actual__lte=F('producto__stock_minimo')).count()
    
    context = {
        'inventario': inventario,
        'total_articulos': inventario.count(),
        'stock_critico_count': stock_critico_count,
    }
    return render(request, 'gestion/inventario_list.html', context)

@login_required
def inventario_historial(request):
    inventario = Inventario.objects.filter(producto__activo=False)
    return render(request, 'gestion/inventario_list.html', {
        'inventario': inventario,
        'es_historial': True,
        'titulo': 'Historial de Productos'
    })

@login_required
def producto_create(request):
    is_restaurante = request.GET.get('categoria') == 'RESTAURANTE' or request.session.get('modulo') == 'RESTAURANTE'
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            if not producto.codigo:
                import random
                producto.codigo = f"prod-{random.randint(1000, 9999)}"
                while Producto.objects.filter(codigo=producto.codigo).exists():
                    producto.codigo = f"prod-{random.randint(1000, 9999)}"
            if is_restaurante:
                producto.modulo = 'RESTAURANTE'
                producto.categoria = 'RESTAURANTE'

            
            # CAPTURA MANUAL DE TODOS LOS STOCKS
            stock_actual = int(request.POST.get('stock', 0) or 0)
            stock_min = int(request.POST.get('stock_minimo', 0) or 0)
            stock_ideal_val = int(request.POST.get('stock_ideal', 0) or 0)
            stock_alert_val = int(request.POST.get('stock_alerta', 0) or 0)
            
            producto.stock_minimo = stock_min
            producto.stock_ideal = stock_ideal_val
            producto.stock_alerta = stock_alert_val
            producto.save()
            form.save_m2m()
                
            # Guardar en Inventario
            inv, created = Inventario.objects.update_or_create(
                producto=producto,
                defaults={'stock_actual': stock_actual}
            )
            
            messages.success(request, f'Producto creado. Stock: {stock_actual}, Ideal: {stock_ideal_val}')
            if is_restaurante:
                return redirect('carta_lista')
            return redirect('inventario_list')
    else:
        # Generar código sugerido (encuentra el número más bajo disponible)
        nuevo_numero = 1
        nuevo_codigo = f"prod-{nuevo_numero:03d}"
        while Producto.objects.filter(codigo__iexact=nuevo_codigo).exists():
            nuevo_numero += 1
            nuevo_codigo = f"prod-{nuevo_numero:03d}"
            
        initial = {'codigo': nuevo_codigo}
        if is_restaurante:
            initial.update({'categoria': 'RESTAURANTE', 'modulo': 'RESTAURANTE'})
        form = ProductoForm(initial=initial)
    
    template = 'gestion/producto_form_restaurante.html' if is_restaurante else 'gestion/producto_form.html'
    return render(request, template, {'form': form, 'titulo': 'Nuevo Artículo'})


@login_required
def producto_update(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    is_restaurante = producto.modulo == 'RESTAURANTE'
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            prod_edit = form.save(commit=False)
            
            # CAPTURA MANUAL DE TODOS LOS STOCKS
            stock_actual = int(request.POST.get('stock', 0) or 0)
            stock_min = int(request.POST.get('stock_minimo', 0) or 0)
            stock_ideal_val = int(request.POST.get('stock_ideal', 0) or 0)
            
            prod_edit.stock_minimo = stock_min
            prod_edit.stock_ideal = stock_ideal_val
            prod_edit.save()
            form.save_m2m()
                
            # Guardar en Inventario
            inv, created = Inventario.objects.update_or_create(
                producto=prod_edit,
                defaults={'stock_actual': stock_actual}
            )
            
            messages.success(request, f'Producto actualizado. Stock: {stock_actual}, Ideal: {stock_ideal_val}')
            if is_restaurante:
                return redirect('carta_lista')
            return redirect('inventario_list')
        else:
            # Mostrar errores si el formulario no es válido
            for field, errors in form.errors.items():
                messages.warning(request, f"Error en {field}: {errors}")
    else:
        form = ProductoForm(instance=producto)
    
    template = 'gestion/producto_form_restaurante.html' if is_restaurante else 'gestion/producto_form.html'
    return render(request, template, {'form': form, 'producto': producto, 'titulo': 'Editar Artículo'})

@login_required
def producto_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.activo = False
    producto.save()
    messages.success(request, "Producto movido al historial.")
    return redirect('inventario_list')

@login_required
def producto_restore(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.activo = True
    producto.save()
    messages.success(request, "Producto restaurado.")
    return redirect('inventario_list')

# --- CATEGORIAS PERSONALIZADAS ---
@login_required
def categorias_list(request):
    modulo = request.session.get('modulo', 'RESTAURANTE')
    categorias = CategoriaPersonalizada.objects.filter(modulo=modulo)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            CategoriaPersonalizada.objects.get_or_create(nombre=nombre, modulo=modulo)
            messages.success(request, "Categoría creada correctamente.")
        return redirect('categorias_list')
        
    return render(request, 'gestion/categorias_list.html', {'categorias': categorias})

@login_required
def categoria_delete(request, pk):
    cat = get_object_or_404(CategoriaPersonalizada, pk=pk)
    cat.delete()
    messages.success(request, "Categoría eliminada.")
    return redirect('categorias_list')

# --- VENTAS ---

@login_required
def ventas_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    ventas = Venta.objects.filter(modulo=modulo).order_by('-fecha')
    
    # Filtro por fecha
    fecha_filtro = request.GET.get('fecha')
    if fecha_filtro:
        ventas = ventas.filter(fecha__date=fecha_filtro)
        
    return render(request, 'gestion/ventas_list.html', {
        'ventas': ventas,
        'fecha_filtro': fecha_filtro
    })

@login_required
def venta_create(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            venta = form.save(commit=False)
            venta.usuario = request.user
            venta.modulo = request.session.get('modulo', 'HOTEL')
            venta.save()
            messages.success(request, "Venta registrada correctamente.")
            return redirect('ventas_list')
    else:
        form = VentaForm()
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Nueva Venta'})

@login_required
def venta_pdf(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    detalles = venta.detalles.all()
    
    # Cálculos para el comprobante
    total = float(venta.total)
    op_gravada = total / 1.18
    igv = total - op_gravada
    
    data = {
        'venta': venta,
        'detalles': detalles,
        'op_gravada': op_gravada,
        'igv': igv,
        'total_letras': "MONTO REFERENCIAL"
    }
    
    pdf = render_to_pdf('gestion/comprobante_pdf.html', data)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"{venta.tipo_comprobante}_{venta.serie}-{venta.numero}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    
    messages.error(request, "No se pudo generar el PDF del comprobante.")
    return redirect('ventas_list')

@login_required
def importar_ventas(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        if not (archivo.name.endswith('.xlsx') or archivo.name.endswith('.xls') or archivo.name.endswith('.csv')):
            messages.error(request, "Solo se admiten archivos .xlsx, .xls o .csv")
            return redirect('ventas_list')

        try:
            ventas_creadas = 0
            if archivo.name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
                rows = ws.iter_rows(min_row=2, values_only=True)
            elif archivo.name.endswith('.xls'):
                import xlrd
                file_bytes = archivo.read()
                try:
                    wb = xlrd.open_workbook(file_contents=file_bytes)
                    sheet = wb.sheet_by_index(0)
                    headers = [str(sheet.cell_value(0, col)).lower().strip() for col in range(sheet.ncols)]
                    rows = []
                    for row_idx in range(1, sheet.nrows):
                        row_vals = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                        rows.append(row_vals)
                except Exception as e_xls:
                    # Si falla, puede que sea un archivo HTML/XML guardado como .xls (común en reportes de SUNAT)
                    if b'<html' in file_bytes.lower() or b'<table' in file_bytes.lower() or b'<?xml' in file_bytes.lower():
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(file_bytes, 'html.parser')
                        table = soup.find('table')
                        if table:
                            tr_list = table.find_all('tr')
                            if tr_list:
                                th_list = tr_list[0].find_all(['th', 'td'])
                                headers = [cell.get_text().lower().strip() for cell in th_list]
                                rows = []
                                for tr in tr_list[1:]:
                                    td_list = tr.find_all(['td', 'th'])
                                    if not td_list:
                                        continue
                                    row_vals = [cell.get_text().strip() for cell in td_list]
                                    if len(row_vals) < len(headers):
                                        row_vals += [""] * (len(headers) - len(row_vals))
                                    rows.append(row_vals)
                            else:
                                raise ValueError("No se encontraron filas en la tabla HTML.")
                        else:
                            raise ValueError("No se encontró una estructura de tabla HTML.")
                    else:
                        raise e_xls

            else:
                decoded_file = archivo.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                headers = [h.lower().strip() for h in next(reader)]
                rows = reader

            # Mapeo de columnas comunes
            idx_fecha = -1
            idx_cliente = -1
            idx_doc = -1
            idx_tipo = -1
            idx_serie = -1
            idx_numero = -1
            idx_total = -1
            idx_condicion = -1   # condicion_vendedor → forma_pago
            idx_estado = -1      # estado → estado_sunat

            for i, h in enumerate(headers):
                if h == 'fecha' or 'emision' in h: idx_fecha = i
                elif h in ('razonsocial', 'razon_social') or 'nombre' in h or 'cliente' in h: idx_cliente = i
                elif h in ('numerodoc', 'nro_doc', 'ruc', 'dni') or 'numerodo' in h: idx_doc = i
                elif h in ('tipodoc', 'desc_tipodoc') or (h == 'tipodoc'): idx_tipo = i
                elif h == 'serie': idx_serie = i
                elif h in ('nro_efact', 'numero', 'correlativo') or ('numero' in h and 'doc' not in h): idx_numero = i
                elif h in ('preciovent', 'total', 'monto', 'importe') or 'precioven' in h: idx_total = i
                elif 'condicion' in h: idx_condicion = i
                elif h == 'estado': idx_estado = i

            if idx_total == -1 or (idx_cliente == -1 and idx_doc == -1):
                messages.error(request, "No se pudieron identificar las columnas necesarias en el archivo.")
                return redirect('ventas_list')

            modulo_actual = request.session.get('modulo', 'HOTEL')

            for row in rows:
                if not any(row): continue
                
                try:
                    fecha_val = row[idx_fecha] if idx_fecha != -1 else None
                    if isinstance(fecha_val, datetime):
                        fecha = fecha_val
                    elif isinstance(fecha_val, (int, float)):
                        try:
                            import xlrd
                            date_tuple = xlrd.xldate_as_tuple(fecha_val, wb.datemode if 'wb' in locals() and hasattr(wb, 'datemode') else 0)
                            fecha = datetime(*date_tuple)
                        except:
                            fecha = timezone.now()
                    elif isinstance(fecha_val, str):
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M'):
                            try:
                                fecha_val_clean = fecha_val.split(' ')[0] if ' ' in fecha_val else fecha_val
                                fecha = datetime.strptime(fecha_val_clean, fmt)
                                break
                            except: continue
                        else: fecha = timezone.now()
                    else: fecha = timezone.now()

                    monto_total = float(row[idx_total]) if row[idx_total] not in ('', None) else 0.0
                    nombre_cliente = str(row[idx_cliente]).strip() if idx_cliente != -1 and row[idx_cliente] not in ('', None) else "Cliente Importado"
                    doc_cliente = str(row[idx_doc]).strip() if idx_doc != -1 and row[idx_doc] not in ('', None) else "00000000"
                    
                    # Limpiar notación científica en DNI/RUC (ej: 2.05E+10)
                    if 'e+' in doc_cliente.lower() or 'e-' in doc_cliente.lower():
                        try:
                            doc_cliente = str(int(float(doc_cliente)))
                        except:
                            pass

                    tipo_raw = str(row[idx_tipo]).upper().strip() if idx_tipo != -1 and row[idx_tipo] not in ('', None) else "BOLETA"
                    if 'FACTURA' in tipo_raw: tipo_comp = 'FACTURA'
                    elif 'NOTA' in tipo_raw and 'CREDITO' in tipo_raw: tipo_comp = 'NOTA_CREDITO'
                    elif 'BOLETA' in tipo_raw: tipo_comp = 'BOLETA'
                    else: tipo_comp = 'BOLETA'
                    
                    serie = str(row[idx_serie]).strip() if idx_serie != -1 and row[idx_serie] not in ('', None) else "E001"
                    numero_val = str(row[idx_numero]).strip() if idx_numero != -1 and row[idx_numero] not in ('', None) else "1"
                    
                    # Limpiar número de factura si contiene serie
                    if idx_serie != -1 and row[idx_serie] not in ('', None):
                        serie_str = str(row[idx_serie]).strip()
                        if numero_val.startswith(serie_str):
                            numero_val = numero_val[len(serie_str):]
                    
                    nro_digits = ''.join(c for c in numero_val if c.isdigit())
                    try: numero = int(nro_digits)
                    except: numero = 1

                    # Forma de pago desde condicion_vendedor
                    if idx_condicion != -1 and row[idx_condicion] not in ('', None):
                        cond_raw = str(row[idx_condicion]).upper().strip()
                        forma_pago = 'CREDITO' if 'CREDIT' in cond_raw else 'CONTADO'
                    else:
                        forma_pago = 'CONTADO'

                    # Estado SUNAT desde columna estado
                    if idx_estado != -1 and row[idx_estado] not in ('', None):
                        est_raw = str(row[idx_estado]).upper().strip()
                        if 'ACEPT' in est_raw: estado_sunat_val = 'ACEPTADO'
                        elif 'RECHAZ' in est_raw: estado_sunat_val = 'RECHAZADO'
                        elif 'ERROR' in est_raw: estado_sunat_val = 'ERROR'
                        else: estado_sunat_val = 'PENDIENTE'
                    else:
                        estado_sunat_val = 'PENDIENTE'

                    cliente, _ = Cliente.objects.get_or_create(
                        numero_documento=doc_cliente,
                        defaults={
                            'nombre_razon_social': nombre_cliente,
                            'tipo_documento': 'RUC' if len(doc_cliente) == 11 else 'DNI'
                        }
                    )

                    Venta.objects.create(
                        cliente=cliente,
                        usuario=request.user,
                        tipo_comprobante=tipo_comp,
                        serie=serie,
                        numero=numero,
                        total=monto_total,
                        forma_pago=forma_pago,
                        estado_sunat=estado_sunat_val,
                        modulo=modulo_actual,
                    )
                    ventas_creadas += 1
                except Exception as e:
                    print(f"Error en fila: {e}")
                    continue

            messages.success(request, f"Importación finalizada: {ventas_creadas} ventas registradas.")
        except Exception as e:
            messages.error(request, f"Error crítico al importar: {str(e)}")
            
    return redirect('ventas_list')

# --- COMPRAS ---

@login_required
def compras_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    compras = Compra.objects.filter(modulo=modulo, activo=True).order_by('-fecha')
    return render(request, 'gestion/compras_list.html', {'compras': compras})

@login_required
def compras_historial(request):
    modulo = request.session.get('modulo', 'HOTEL')
    compras = Compra.objects.filter(modulo=modulo, activo=False).order_by('-fecha')
    return render(request, 'gestion/compras_list.html', {
        'compras': compras,
        'es_historial': True,
        'titulo': 'Historial de Compras (Eliminadas)'
    })

@login_required
def compra_create(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                compra = form.save(commit=False)
                compra.modulo = request.session.get('modulo', 'HOTEL')
                compra.save()
                
                # Guardar Detalles (Items)
                productos = request.POST.getlist('producto[]')
                cantidades = request.POST.getlist('cantidad[]')
                precios = request.POST.getlist('precio[]')
                
                for i in range(len(productos)):
                    if productos[i]:
                        DetalleCompra.objects.create(
                            compra=compra,
                            producto_id=productos[i],
                            cantidad=cantidades[i],
                            precio_unitario=precios[i],
                            subtotal=float(cantidades[i]) * float(precios[i])
                        )
                        # Actualizar Inventario
                        inventario, _ = Inventario.objects.get_or_create(producto_id=productos[i])
                        inventario.stock_actual += int(cantidades[i])
                        inventario.save()
                
                # Guardar Cuotas si es crédito
                if compra.forma_pago == 'CREDITO':
                    num_cuotas = int(request.POST.get('num_cuotas', 1))
                    compra.num_cuotas = num_cuotas
                    compra.save()
                    
                    montos_cuotas = request.POST.getlist('monto_cuota[]')
                    fechas_cuotas = request.POST.getlist('fecha_cuota[]')
                    
                    for i in range(num_cuotas):
                        CuotaCompra.objects.create(
                            compra=compra,
                            numero_cuota=i+1,
                            monto=montos_cuotas[i],
                            fecha_vencimiento=fechas_cuotas[i]
                        )
                
                messages.success(request, "Compra registrada con éxito y stock actualizado.")
                return redirect('compras_list')
    else:
        form = CompraForm()
    
    productos = Producto.objects.filter(activo=True)
    return render(request, 'gestion/compra_form.html', {
        'form': form, 
        'productos': productos,
        'titulo': 'Registrar Compra Detallada'
    })

@login_required
def compra_detalle(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    detalles = compra.detalles.all()
    cuotas = compra.cuotas.all().order_by('numero_cuota')
    return render(request, 'gestion/compra_detalle.html', {
        'compra': compra,
        'detalles': detalles,
        'cuotas': cuotas,
        'titulo': f'Detalle de Compra {compra.numero_comprobante}'
    })

@login_required
def compra_delete(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    compra.activo = False
    compra.save()
    messages.success(request, "Compra movida al historial.")
    return redirect('compras_list')

@login_required
def compra_restore(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    compra.activo = True
    compra.save()
    messages.success(request, "Compra restaurada correctamente.")
    return redirect('compras_list')

# --- SERVICIOS ---

@login_required
def servicios_list(request):
    modulo = request.session.get('modulo', 'HOTEL')
    servicios = RegistroServicio.objects.filter(modulo=modulo).order_by('-fecha')
    catalogo = Producto.objects.filter(categoria__in=['TURISMO', 'HOSPEDAJE'], activo=True).order_by('nombre')
    
    return render(request, 'gestion/servicios_list.html', {
        'servicios': servicios,
        'catalogo': catalogo,
        'titulo': 'Gestión de Servicios'
    })

@login_required
def servicio_create(request):
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            servicio = form.save(commit=False)
            servicio.modulo = request.session.get('modulo', 'HOTEL')
            servicio.save()
            messages.success(request, "Servicio registrado.")
            return redirect('servicios_list')
    else:
        form = ServicioForm()
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Nuevo Registro de Servicio'})

@login_required
def servicio_update(request, pk):
    servicio = get_object_or_404(RegistroServicio, pk=pk)
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, "Servicio actualizado.")
            return redirect('servicios_list')
    else:
        form = ServicioForm(instance=servicio)
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Editar Servicio'})

@login_required
def servicio_delete(request, pk):
    servicio = get_object_or_404(RegistroServicio, pk=pk)
    servicio.delete()
    messages.success(request, "Servicio eliminado.")
    return redirect('servicios_list')

# --- FACTURACION / SUNAT ---

@login_required
def facturacion_list(request):
    facturas = FacturaElectronica.objects.all().order_by('-fecha_emision')
    return render(request, 'gestion/facturacion_list.html', {'facturas': facturas})

@login_required
def enviar_sunat(request, pk):
    # Simulación de envío a SUNAT
    messages.success(request, f"Comprobante #{pk} enviado correctamente a SUNAT.")
    return redirect('facturacion_list')

# --- REPORTES / EXCEL ---

@login_required
def reportes_view(request):
    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario')
    periodo = request.GET.get('periodo', 'mes') # dia, semana, mes
    
    # 1. Ventas QuerySet
    ventas_qs = Venta.objects.all()
    servicios_qs = RegistroServicio.objects.all()
    
    if fecha_inicio:
        ventas_qs = ventas_qs.filter(fecha__date__gte=fecha_inicio)
        servicios_qs = servicios_qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        ventas_qs = ventas_qs.filter(fecha__date__lte=fecha_fin)
        servicios_qs = servicios_qs.filter(fecha__date__lte=fecha_fin)
    if usuario_id and usuario_id != 'Todos':
        ventas_qs = ventas_qs.filter(usuario_id=usuario_id)
        
    total_ventas_monto = ventas_qs.aggregate(Sum('total'))['total__sum'] or 0
    cantidad_ventas = ventas_qs.count()

    # --- AGRUPACIÓN DE VENTAS ---
    from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
    
    if periodo == 'dia':
        trunc_func = TruncDay('fecha')
        date_format = '%d/%m'
    elif periodo == 'semana':
        trunc_func = TruncWeek('fecha')
        date_format = 'Sem %W'
    else: # mes
        trunc_func = TruncMonth('fecha')
        date_format = '%b'

    ventas_agrupadas = ventas_qs.annotate(
        time_unit=trunc_func
    ).values('time_unit').annotate(total=Sum('total')).order_by('time_unit')
    
    sales_labels = []
    sales_data = []
    
    for v in ventas_agrupadas:
        if v['time_unit']:
            label = v['time_unit'].strftime(date_format)
            sales_labels.append(label)
            sales_data.append(float(v['total']))
        
    # --- SERVICIOS POPULARES ---
    servicios_query = servicios_qs.values('nombre_servicio').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    services_labels = [s['nombre_servicio'] for s in servicios_query]
    services_data = [s['count'] for s in servicios_query]
    
    # --- TOP CLIENTES ---
    clientes_top = Cliente.objects.filter(venta__in=ventas_qs).distinct().annotate(
        num_ventas=Count('venta'),
        total_invertido=Sum('venta__total')
    ).order_by('-total_invertido')[:10]
    
    # 2. Compras QuerySet
    compras_qs = Compra.objects.filter(activo=True)
    if fecha_inicio:
        compras_qs = compras_qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        compras_qs = compras_qs.filter(fecha__lte=fecha_fin)
        
    total_compras_monto = compras_qs.aggregate(Sum('total'))['total__sum'] or 0
    cantidad_compras = compras_qs.count()

    # 3. Reporte de Usuarios (Vendedores)
    usuarios_report = []
    for u in User.objects.all():
        ventas_u = Venta.objects.filter(usuario=u)
        if fecha_inicio:
            ventas_u = ventas_u.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            ventas_u = ventas_u.filter(fecha__date__lte=fecha_fin)
        tot_val = ventas_u.aggregate(Sum('total'))['total__sum'] or 0
        cnt_val = ventas_u.count()
        if cnt_val > 0 or tot_val > 0:
            usuarios_report.append({
                'user': u,
                'num_ventas': cnt_val,
                'total_vendido': tot_val,
                'promedio_venta': float(tot_val / cnt_val) if cnt_val > 0 else 0.0
            })
    usuarios_report.sort(key=lambda x: x['total_vendido'], reverse=True)

    # 4. Reporte de Inventario Completo
    inventario_report = Inventario.objects.select_related('producto').filter(producto__activo=True).order_by('producto__subcategoria', 'producto__nombre')
    total_stock_items = inventario_report.aggregate(Sum('stock_actual'))['stock_actual__sum'] or 0
    total_valoracion = sum(float(item.stock_actual) * float(item.producto.precio_venta) for item in inventario_report)

    # 5. Reporte Restaurante
    ventas_rest = Venta.objects.filter(modulo='RESTAURANTE', es_comanda=False)
    if fecha_inicio:
        ventas_rest = ventas_rest.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        ventas_rest = ventas_rest.filter(fecha__date__lte=fecha_fin)
    total_rest = ventas_rest.aggregate(Sum('total'))['total__sum'] or 0
    total_rest_tarjeta = ventas_rest.filter(pago_tarjeta=True).aggregate(Sum('recargo_tarjeta'))['recargo_tarjeta__sum'] or 0
    # Platos más vendidos
    platos_mas_vendidos = DetalleVenta.objects.filter(
        venta__in=ventas_rest,
        producto__categoria='RESTAURANTE'
    ).values('producto__nombre', 'producto__subcategoria').annotate(
        total_vendido=Sum('cantidad'),
        total_ingresos=Sum('subtotal')
    ).order_by('-total_vendido')[:10]
    # Ventas por mozo
    ventas_por_mozo = DetalleVenta.objects.filter(
        venta__in=ventas_rest
    ).exclude(mozo=None).values(
        'mozo__username', 'mozo__first_name', 'mozo__last_name'
    ).annotate(
        total_items=Sum('cantidad'),
        total_ingresos=Sum('subtotal')
    ).order_by('-total_ingresos')
    # Comisiones
    comisiones_por_mozo = DetalleVenta.objects.filter(
        venta__in=ventas_rest
    ).exclude(mozo=None).select_related('mozo', 'producto')
    comision_dict = {}
    for d in comisiones_por_mozo:
        mozo_key = d.mozo.username
        if mozo_key not in comision_dict:
            comision_dict[mozo_key] = {'nombre': d.mozo.get_full_name() or d.mozo.username, 'total_comision': 0}
        comision_dict[mozo_key]['total_comision'] += float(d.producto.comision) * d.cantidad
    comisiones_list = sorted(comision_dict.values(), key=lambda x: x['total_comision'], reverse=True)
    # Ventas incluidas en tour
    detalles_tour_report = DetalleVenta.objects.filter(
        venta__in=ventas_rest, incluido_tour=True
    ).select_related('venta', 'venta__mesa', 'producto').order_by('-venta__fecha')
    
    for d in detalles_tour_report:
        d.costo_real = d.cantidad * (d.producto.precio_venta or Decimal('0.00'))
        
    total_costo_real_tour = sum(d.costo_real for d in detalles_tour_report)
    total_real_obtenido = (total_rest or Decimal('0.00')) + total_costo_real_tour - (total_rest_tarjeta or Decimal('0.00'))
    items_incluidos_tour = detalles_tour_report.count()


    # Guías con más comanda
    guias_report = ventas_rest.exclude(encargado_guia=None).exclude(encargado_guia='').values('encargado_guia').annotate(
        total_ventas=Count('id'),
        total_monto=Sum('total')
    ).order_by('-total_monto')[:10]
    
    context = {

        # Ventas
        'sales_labels': json.dumps(sales_labels),
        'sales_data': json.dumps(sales_data),
        'services_labels': json.dumps(services_labels),
        'services_data': json.dumps(services_data),
        'clientes_top': clientes_top,
        'ventas_list': ventas_qs.order_by('-fecha')[:100],
        'total_ventas_monto': total_ventas_monto,
        'cantidad_ventas': cantidad_ventas,
        
        # Compras
        'compras_list': compras_qs.order_by('-fecha')[:100],
        'total_compras_monto': total_compras_monto,
        'cantidad_compras': cantidad_compras,
        
        # Usuarios
        'usuarios_report': usuarios_report,
        
        # Inventario
        'inventario_report': inventario_report,
        'total_stock_items': total_stock_items,
        'total_valoracion': total_valoracion,

        # Restaurante
        'ventas_rest': ventas_rest.order_by('-fecha')[:100],
        'total_rest': total_rest,
        'total_rest_tarjeta': total_rest_tarjeta,
        'platos_mas_vendidos': platos_mas_vendidos,
        'ventas_por_mozo': ventas_por_mozo,
        'comisiones_list': comisiones_list,
        'items_incluidos_tour': items_incluidos_tour,
        'guias_report': guias_report,
        'detalles_tour_report': detalles_tour_report,
        'total_costo_real_tour': total_costo_real_tour,
        'total_real_obtenido': total_real_obtenido,

        
        # Filtros
        'usuarios': User.objects.all(),
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'usuario_seleccionado': usuario_id,
        'periodo_seleccionado': periodo,
    }
    return render(request, 'gestion/reportes.html', context)


def _insertar_cabecera_y_logo_excel(ws, titulo, num_columnas=8):
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from .models import ConfiguracionEmpresa
    import os
    
    # Insertar 4 filas en blanco al principio
    ws.insert_rows(1, 4)
    
    # Cargar config
    config = ConfiguracionEmpresa.objects.first()
    
    # 1. Agregar logo en A1 si existe
    if config and config.logo and os.path.exists(config.logo.path):
        try:
            img = OpenpyxlImage(config.logo.path)
            # Redimensionar logo a un tamaño moderado para Excel
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception:
            pass
            
    # 2. Datos de la empresa
    company_font = Font(name='Calibri', size=11, bold=True)
    info_font = Font(name='Calibri', size=9, color='555555')
    if config:
        ws['B1'] = config.nombre_comercial or config.razon_social or "TURISMO ERP"
        ws['B1'].font = company_font
        ws['B2'] = f"RUC: {config.ruc or '20123456789'}"
        ws['B2'].font = info_font
        ws['B3'] = f"Dirección: {config.direccion or ''}"
        ws['B3'].font = info_font
    else:
        ws['B1'] = "TURISMO ERP"
        ws['B1'].font = company_font
        ws['B2'] = "RUC: 20123456789"
        ws['B2'].font = info_font
        
    # 3. Título del Reporte en la fila 4
    col_letra_fin = chr(64 + max(2, min(num_columnas, 26)))
    ws.merge_cells(f"A4:{col_letra_fin}4")
    ws['A4'] = titulo.upper()
    ws['A4'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 22


@login_required
def exportar_ventas_generales_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    hotel_ventas = Venta.objects.filter(modulo='HOTEL')
    restaurante_ventas = Venta.objects.filter(modulo='RESTAURANTE')
    
    if fecha_inicio:
        hotel_ventas = hotel_ventas.filter(fecha__date__gte=fecha_inicio)
        restaurante_ventas = restaurante_ventas.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        hotel_ventas = hotel_ventas.filter(fecha__date__lte=fecha_fin)
        restaurante_ventas = restaurante_ventas.filter(fecha__date__lte=fecha_fin)
        
    hotel_ventas = hotel_ventas.order_by('fecha')
    restaurante_ventas = restaurante_ventas.order_by('fecha')
    
    total_hotel = hotel_ventas.aggregate(Sum('total'))['total__sum'] or 0
    total_restaurante = restaurante_ventas.aggregate(Sum('total'))['total__sum'] or 0
    total_general = total_hotel + total_restaurante
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ventas_generales_hotel_restaurante.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas Generales"
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=13, bold=True, color='000000')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    hotel_header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    rest_header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    summary_header_fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    zebra_fill = PatternFill(start_color='F2F4F4', end_color='F2F4F4', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title Block
    ws.merge_cells('A1:I2')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE VENTAS GENERALES (HOTEL Y RESTAURANTE)"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = bold_font
    if fecha_inicio or fecha_fin:
        ws['A4'] = f"Rango: {fecha_inicio or 'Inicio'} hasta {fecha_fin or 'Fin'}"
        ws['A4'].font = bold_font
        
    # --- TABLE 1: HOTEL SALES ---
    ws['A6'] = "CUADRO 1: VENTAS DE HOTEL"
    ws['A6'].font = section_font
    
    headers = ['ID Venta', 'Fecha', 'Cliente', 'Tipo Comprobante', 'Serie-Número', 'Total (S/)', 'Vendedor', 'Forma Pago', 'Estado SUNAT']
    
    row_num = 8
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = hotel_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    row_num += 1
    start_row_hotel = row_num
    for idx, v in enumerate(hotel_ventas):
        ws.cell(row=row_num, column=1, value=v.id).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=v.fecha.strftime('%d/%m/%Y %H:%M')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=v.cliente.nombre_razon_social)
        ws.cell(row=row_num, column=4, value=v.tipo_comprobante).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5, value=f"{v.serie}-{v.numero}").alignment = Alignment(horizontal='center')
        
        c_tot = ws.cell(row=row_num, column=6, value=float(v.total))
        c_tot.number_format = 'S/ #,##0.00'
        c_tot.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=7, value=v.usuario.username)
        ws.cell(row=row_num, column=8, value=v.forma_pago).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=9, value=v.estado_sunat).alignment = Alignment(horizontal='center')
        
        for col_idx in range(1, 10):
            c = ws.cell(row=row_num, column=col_idx)
            c.font = normal_font
            c.border = thin_border
            if idx % 2 == 1:
                c.fill = zebra_fill
        row_num += 1
        
    end_row_hotel = row_num - 1
    # Total Hotel Row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    tot_label = ws.cell(row=row_num, column=1, value="TOTAL VENTAS HOTEL:")
    tot_label.font = bold_font
    tot_label.alignment = Alignment(horizontal='right')
    
    tot_val = ws.cell(row=row_num, column=6)
    if hotel_ventas.exists():
        tot_val.value = f"=SUM(F{start_row_hotel}:F{end_row_hotel})"
    else:
        tot_val.value = 0.00
    tot_val.font = Font(name='Calibri', size=11, bold=True, color='0000FF')
    tot_val.number_format = 'S/ #,##0.00'
    tot_val.border = thin_border
    
    for col_idx in range(1, 10):
        ws.cell(row=row_num, column=col_idx).border = thin_border
        
    # --- TABLE 2: RESTAURANTE SALES ---
    row_num += 3
    ws.cell(row=row_num, column=1, value="CUADRO 2: VENTAS DE RESTAURANTE").font = section_font
    
    row_num += 2
    for col_idx, header in enumerate(headers[:-1], 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = rest_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    row_num += 1
    start_row_rest = row_num
    for idx, v in enumerate(restaurante_ventas):
        ws.cell(row=row_num, column=1, value=v.id).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=v.fecha.strftime('%d/%m/%Y %H:%M')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=v.cliente.nombre_razon_social)
        ws.cell(row=row_num, column=4, value=v.tipo_comprobante).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5, value=f"{v.serie}-{v.numero}").alignment = Alignment(horizontal='center')
        
        c_tot = ws.cell(row=row_num, column=6, value=float(v.total))
        c_tot.number_format = 'S/ #,##0.00'
        c_tot.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=7, value=v.usuario.username)
        ws.cell(row=row_num, column=8, value=v.forma_pago).alignment = Alignment(horizontal='center')
        
        for col_idx in range(1, 9):
            c = ws.cell(row=row_num, column=col_idx)
            c.font = normal_font
            c.border = thin_border
            if idx % 2 == 1:
                c.fill = zebra_fill
        row_num += 1
        
    end_row_rest = row_num - 1
    # Total Restaurante Row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    tot_label = ws.cell(row=row_num, column=1, value="TOTAL VENTAS RESTAURANTE:")
    tot_label.font = bold_font
    tot_label.alignment = Alignment(horizontal='right')
    
    tot_val = ws.cell(row=row_num, column=6)
    if restaurante_ventas.exists():
        tot_val.value = f"=SUM(F{start_row_rest}:F{end_row_rest})"
    else:
        tot_val.value = 0.00
    tot_val.font = Font(name='Calibri', size=11, bold=True, color='27AE60')
    tot_val.number_format = 'S/ #,##0.00'
    tot_val.border = thin_border
    
    for col_idx in range(1, 9):
        ws.cell(row=row_num, column=col_idx).border = thin_border
        
    # --- SUMMARY TABLE ---
    row_num += 3
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    sum_header = ws.cell(row=row_num, column=1, value="RESUMEN GENERAL DE INGRESOS")
    sum_header.font = header_font
    sum_header.fill = summary_header_fill
    sum_header.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="Concepto").font = bold_font
    ws.cell(row=row_num, column=1).border = thin_border
    ws.cell(row=row_num, column=2, value="Total Acumulado").font = bold_font
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=2).border = thin_border
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="Ventas Hotel").font = normal_font
    ws.cell(row=row_num, column=1).border = thin_border
    c_sh = ws.cell(row=row_num, column=2, value=float(total_hotel))
    c_sh.number_format = 'S/ #,##0.00'
    c_sh.alignment = Alignment(horizontal='right')
    c_sh.border = thin_border
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="Ventas Restaurante").font = normal_font
    ws.cell(row=row_num, column=1).border = thin_border
    c_sr = ws.cell(row=row_num, column=2, value=float(total_restaurante))
    c_sr.number_format = 'S/ #,##0.00'
    c_sr.alignment = Alignment(horizontal='right')
    c_sr.border = thin_border
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTAL GENERAL").font = bold_font
    ws.cell(row=row_num, column=1).border = thin_border
    c_sg = ws.cell(row=row_num, column=2, value=float(total_general))
    c_sg.font = bold_font
    c_sg.number_format = 'S/ #,##0.00'
    c_sg.alignment = Alignment(horizontal='right')
    c_sg.border = thin_border
    
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in ws.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    _insertar_cabecera_y_logo_excel(ws, "Reporte General de Ventas (Hotel y Restaurante)", 9)
    wb.save(response)
    return response


@login_required
def exportar_ventas_generales_pdf(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    hotel_ventas = Venta.objects.filter(modulo='HOTEL')
    restaurante_ventas = Venta.objects.filter(modulo='RESTAURANTE')
    
    if fecha_inicio:
        hotel_ventas = hotel_ventas.filter(fecha__date__gte=fecha_inicio)
        restaurante_ventas = restaurante_ventas.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        hotel_ventas = hotel_ventas.filter(fecha__date__lte=fecha_fin)
        restaurante_ventas = restaurante_ventas.filter(fecha__date__lte=fecha_fin)
        
    hotel_ventas = hotel_ventas.order_by('fecha')
    restaurante_ventas = restaurante_ventas.order_by('fecha')
    
    total_hotel = hotel_ventas.aggregate(Sum('total'))['total__sum'] or 0
    total_restaurante = restaurante_ventas.aggregate(Sum('total'))['total__sum'] or 0
    total_general = total_hotel + total_restaurante
    
    empresa_config = ConfiguracionEmpresa.objects.first()
    
    data = {
        'fecha': datetime.now(),
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'hotel_ventas': hotel_ventas,
        'restaurante_ventas': restaurante_ventas,
        'total_hotel': total_hotel,
        'total_restaurante': total_restaurante,
        'total_general': total_general,
        'empresa_config': empresa_config,
    }
    
    pdf = render_to_pdf('gestion/reporte_ventas_generales_pdf.html', data)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Reporte_Ventas_Generales.pdf"'
        return response
    return HttpResponse("Error generando reporte", status=400)


@login_required
def exportar_compras_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    compras = Compra.objects.filter(activo=True).order_by('fecha')
    if fecha_inicio:
        compras = compras.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        compras = compras.filter(fecha__lte=fecha_fin)
        
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte_compras.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    
    columns = ['ID Compra', 'Fecha', 'Proveedor', 'RUC Proveedor', 'Tipo Comprobante', 'Número Comprobante', 'Forma de Pago', 'N° Cuotas', 'Total (S/)']
    ws.append(columns)
    
    for c in compras:
        ws.append([
            c.id,
            c.fecha.strftime('%d/%m/%Y'),
            c.proveedor.razon_social,
            c.proveedor.ruc,
            c.tipo_comprobante,
            c.numero_comprobante,
            c.forma_pago,
            c.num_cuotas,
            float(c.total)
        ])
    
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    _insertar_cabecera_y_logo_excel(ws, "Reporte de Compras Realizadas", 9)
    wb.save(response)
    return response


@login_required
def exportar_usuarios_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ventas_por_usuario.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas por Usuario"
    
    columns = ['ID Usuario', 'Usuario', 'Nombre Completo', 'Cantidad Ventas', 'Total Vendido (S/)', 'Promedio Venta (S/)']
    ws.append(columns)
    
    for u in User.objects.all():
        ventas_u = Venta.objects.filter(usuario=u)
        if fecha_inicio:
            ventas_u = ventas_u.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            ventas_u = ventas_u.filter(fecha__date__lte=fecha_fin)
        tot_val = ventas_u.aggregate(Sum('total'))['total__sum'] or 0
        cnt_val = ventas_u.count()
        if cnt_val > 0 or tot_val > 0:
            ws.append([
                u.id,
                u.username,
                u.get_full_name() or u.username,
                cnt_val,
                float(tot_val),
                float(tot_val / cnt_val) if cnt_val > 0 else 0.0
            ])
            
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    _insertar_cabecera_y_logo_excel(ws, "Reporte de Ventas por Usuario", 6)
    wb.save(response)
    return response


@login_required
def exportar_inventario_excel(request):
    inventario = Inventario.objects.select_related('producto').filter(producto__activo=True).order_by('producto__categoria', 'producto__nombre')
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="inventario_actual.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    columns = ['Código', 'Producto', 'Categoría', 'Subcategoría', 'Módulo', 'Stock Actual', 'Stock Mínimo', 'Stock Ideal', 'Precio Compra (S/)', 'Precio Venta (S/)', 'Valorización Stock (S/)']
    ws.append(columns)
    
    for item in inventario:
        val = float(item.stock_actual) * float(item.producto.precio_venta)
        ws.append([
            item.producto.codigo,
            item.producto.nombre,
            item.producto.get_categoria_display(),
            item.producto.subcategoria or '',
            item.producto.modulo,
            item.stock_actual,
            item.producto.stock_minimo,
            item.producto.stock_ideal,
            float(item.producto.precio_compra),
            float(item.producto.precio_venta),
            val
        ])
        
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    _insertar_cabecera_y_logo_excel(ws, "Reporte de Inventario Actual", 11)
    wb.save(response)
    return response

@login_required
def exportar_ventas_excel(request):
    modulo = request.session.get('modulo', 'HOTEL')
    fecha_filtro = request.GET.get('fecha')
    
    ventas = Venta.objects.filter(modulo=modulo).order_by('fecha')
    if fecha_filtro:
        ventas = ventas.filter(fecha__date=fecha_filtro)
        
    filename = f"ventas_{modulo.lower()}.xlsx"
    if fecha_filtro:
        filename = f"ventas_{modulo.lower()}_{fecha_filtro}.xlsx"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas {modulo.capitalize()}"
    columns = ['ID', 'Fecha', 'Cliente', 'Tipo', 'Serie-Numero', 'Total', 'Forma Pago', 'Estado SUNAT']
    ws.append(columns)
    
    for v in ventas:
        ws.append([v.id, v.fecha.strftime('%d/%m/%Y %H:%M'), v.cliente.nombre_razon_social,
                   v.tipo_comprobante, f"{v.serie}-{v.numero}", float(v.total), v.forma_pago, v.estado_sunat])
    wb.save(response)
    return response

@login_required
def reporte_general_pdf(request):
    today = timezone.now()
    month_start = today.replace(day=1, hour=0, minute=0, second=0)
    
    modulo = request.session.get('modulo', 'HOTEL')
    
    # Estadísticas para el reporte
    ingresos_mes = Venta.objects.filter(fecha__gte=month_start, modulo=modulo).aggregate(Sum('total'))['total__sum'] or 0
    servicios_totales = RegistroServicio.objects.count() if modulo == 'HOTEL' else 0
    nuevos_clientes = Cliente.objects.filter(id__gt=0).count() # General
    stock_critico = Inventario.objects.filter(producto__modulo=modulo, stock_actual__lte=F('producto__stock_minimo')).count()
    
    # Top Clientes (con suma de ventas filtrado por modulo)
    clientes_top = Cliente.objects.filter(venta__modulo=modulo).annotate(
        num_ventas=Count('venta'),
        total_invertido=Sum('venta__total')
    ).order_by('-total_invertido')[:5]
    
    data = {
        'modulo': modulo,
        'fecha': today,
        'ingresos_mes': ingresos_mes,
        'servicios_totales': servicios_totales,
        'nuevos_clientes': nuevos_clientes,
        'stock_critico': stock_critico,
        'clientes_top': clientes_top,
        'ventas_recientes': Venta.objects.filter(modulo=modulo).order_by('-fecha')[:10]
    }
    
    pdf = render_to_pdf('gestion/reporte_pro_pdf.html', data)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Reporte_Gestion_TurismoERP.pdf"'
        return response
    return HttpResponse("Error generando reporte", status=400)

@login_required
def configuracion_empresa(request):
    config, created = ConfiguracionEmpresa.objects.get_or_create(id=1, defaults={
        'ruc': '20123456789',
        'razon_social': 'Mi Empresa Turística S.A.C.',
        'direccion': 'Calle Principal 123, Arequipa',
    })
    
    if request.method == 'POST':
        form = ConfiguracionEmpresaForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración guardada correctamente.")
            return redirect('configuracion_empresa')
    else:
        form = ConfiguracionEmpresaForm(instance=config)
        
    return render(request, 'gestion/configuracion.html', {'form': form, 'config': config})

# --- HOTEL: HABITACIONES Y RESERVAS ---

@login_required
def habitaciones_list(request):
    habitaciones = Habitacion.objects.filter(activo=True).order_by('numero')

    # Filtros
    filtro_tipo = request.GET.get('tipo', '')
    filtro_estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if filtro_tipo:
        habitaciones = habitaciones.filter(tipo=filtro_tipo)
    if filtro_estado:
        habitaciones = habitaciones.filter(estado=filtro_estado)

    # Filtro de disponibilidad por rango de fechas
    if fecha_desde and fecha_hasta:
        try:
            from datetime import datetime
            d_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            d_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            # Habitaciones ocupadas en ese rango
            hab_ocupadas_ids = Reserva.objects.filter(
                activo=True,
                estado='CHECKIN',
                fecha_ingreso__lt=d_hasta,
                fecha_salida__gt=d_desde
            ).values_list('habitacion_id', flat=True)
            habitaciones = habitaciones.exclude(id__in=hab_ocupadas_ids)
        except ValueError:
            pass

    # Obtener la reserva activa para cada habitación ocupada o reservada
    for hab in habitaciones:
        if hab.estado == 'OCUPADA':
            hab.reserva_activa = Reserva.objects.filter(
                habitacion=hab, estado='CHECKIN', activo=True
            ).last()
        elif hab.estado == 'RESERVADA':
            hab.reserva_activa = Reserva.objects.filter(
                habitacion=hab, estado='PENDIENTE', activo=True
            ).last()

    context = {
        'habitaciones': habitaciones,
        'filtro_tipo': filtro_tipo,
        'filtro_estado': filtro_estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos_habitacion': Habitacion.TIPO_CHOICES,
        'estados_habitacion': Habitacion.ESTADO_CHOICES,
        'personales_limpieza': User.objects.filter(perfil__rol='LIMPIEZA'),
    }
    return render(request, 'gestion/habitaciones_list.html', context)

@login_required
def habitaciones_historial(request):
    habitaciones = Habitacion.objects.filter(activo=False).order_by('numero')
    return render(request, 'gestion/habitaciones_list.html', {
        'habitaciones': habitaciones,
        'es_historial': True,
        'titulo': 'Historial de Habitaciones (Eliminadas)'
    })


@login_required
def habitacion_create(request):
    if request.method == 'POST':
        form = HabitacionForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Habitación registrada.")
            return redirect('habitaciones_list')
    else:
        form = HabitacionForm()
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Nueva Habitación'})

@login_required
def habitacion_update(request, pk):
    habitacion = get_object_or_404(Habitacion, pk=pk)
    if request.method == 'POST':
        form = HabitacionForm(request.POST, request.FILES, instance=habitacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Habitación actualizada correctamente.")
            return redirect('habitaciones_list')
    else:
        form = HabitacionForm(instance=habitacion)
    return render(request, 'gestion/base_form.html', {'form': form, 'titulo': 'Editar Habitación'})

@login_required
def habitacion_delete(request, pk):
    habitacion = get_object_or_404(Habitacion, pk=pk)
    habitacion.activo = False
    habitacion.save()
    messages.success(request, "Habitación movida al historial.")
    return redirect('habitaciones_list')

@login_required
def habitacion_restore(request, pk):
    habitacion = get_object_or_404(Habitacion, pk=pk)
    habitacion.activo = True
    habitacion.save()
    messages.success(request, "Habitación restaurada correctamente.")
    return redirect('habitaciones_list')

@login_required
def reserva_list(request):
    from django.db.models import Q
    reservas = Reserva.objects.filter(activo=True)
    
    # Filtro de búsqueda de texto
    query = request.GET.get('q', '').strip()
    if query:
        reservas = reservas.filter(
            Q(cliente__nombre_razon_social__icontains=query) |
            Q(habitacion__numero__icontains=query) |
            Q(placa_vehiculo__icontains=query)
        )
        
    # Filtro de estado
    estado_filtro = request.GET.get('estado', '').strip()
    if estado_filtro:
        reservas = reservas.filter(estado=estado_filtro)
        
    reservas = reservas.order_by('-fecha_creacion')
    return render(request, 'gestion/reservas_list.html', {
        'reservas': reservas,
        'query': query,
        'estado_filtro': estado_filtro
    })

@login_required
def reservas_historial(request):
    from django.db.models import Q
    reservas = Reserva.objects.filter(activo=False)
    
    # Filtro de búsqueda de texto
    query = request.GET.get('q', '').strip()
    if query:
        reservas = reservas.filter(
            Q(cliente__nombre_razon_social__icontains=query) |
            Q(habitacion__numero__icontains=query) |
            Q(placa_vehiculo__icontains=query)
        )
        
    # Filtro de estado
    estado_filtro = request.GET.get('estado', '').strip()
    if estado_filtro:
        reservas = reservas.filter(estado=estado_filtro)
        
    reservas = reservas.order_by('-fecha_creacion')
    return render(request, 'gestion/reservas_list.html', {
        'reservas': reservas,
        'es_historial': True,
        'titulo': 'Historial de Recepción (Eliminados)',
        'query': query,
        'estado_filtro': estado_filtro
    })

@login_required
def reserva_create(request):
    habitacion_id = request.GET.get('hab')
    if request.method == 'POST':
        form = ReservaForm(request.POST)
        if form.is_valid():
            reserva = form.save(commit=False)
            tipo_accion = request.POST.get('tipo_accion', 'CHECKIN')  # 'RESERVAR' o 'CHECKIN'
            
            # Verificar si se ingresó un nombre rápido de cliente
            nombre_rapido = request.POST.get('nombre_cliente_rapido', '').strip()
            if not form.cleaned_data.get('cliente') and not nombre_rapido:
                form.add_error('cliente', 'Debe seleccionar un cliente registrado o ingresar un nombre de cliente rápido.')
            else:
                if nombre_rapido and not form.cleaned_data.get('cliente'):
                    import uuid
                    temp_doc = f"TEMP-{uuid.uuid4().hex[:8].upper()}"
                    cliente = Cliente.objects.create(
                        tipo_documento='DNI',
                        numero_documento=temp_doc,
                        nombre_razon_social=nombre_rapido,
                        modulo='HOTEL',
                        activo=True
                    )
                    reserva.cliente = cliente

                if tipo_accion == 'RESERVAR':
                    reserva.estado = 'PENDIENTE'
                    reserva.save()
                    reserva.habitacion.estado = 'RESERVADA'
                    reserva.habitacion.save()
                    messages.success(request, f"Reserva registrada para Hab. {reserva.habitacion.numero}. Pendiente de Check-in.")
                else:
                    reserva.estado = 'CHECKIN'
                    reserva.save()
                    reserva.habitacion.estado = 'OCUPADA'
                    reserva.habitacion.save()
                    messages.success(request, "Check-in realizado correctamente.")
                return redirect('reserva_list')
    else:
        initial_data = {}
        if habitacion_id:
            initial_data['habitacion'] = habitacion_id
        form = ReservaForm(initial=initial_data)
    
    habitaciones_precios = {h.id: float(h.precio_noche) for h in Habitacion.objects.all()}
    
    return render(request, 'gestion/reserva_form.html', {
        'form': form, 
        'titulo': 'Nueva Reserva / Check-in',
        'precios_json': json.dumps(habitaciones_precios)
    })

@login_required
def pedido_habitacion_create(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        form = PedidoHabitacionForm(request.POST)
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.reserva = reserva
            pedido.subtotal = pedido.cantidad * pedido.precio_unitario
            pedido.save()
            messages.success(request, "Pedido registrado a la habitación.")
            return redirect('reserva_list')
    else:
        form = PedidoHabitacionForm()
    
    # Preparamos la información de los productos para el JS
    productos_info = {
        p.id: {
            'precio': float(p.precio_venta),
            'categoria': p.categoria,
            'categoria_display': p.get_categoria_display()
        } for p in Producto.objects.filter(activo=True)
    }
    
    return render(request, 'gestion/pedido_form.html', {
        'form': form, 
        'titulo': f'Nuevo Pedido - Hab. {reserva.habitacion.numero}',
        'productos_info_json': json.dumps(productos_info),
        'reserva': reserva
    })

@login_required
def reserva_checkout(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    pedidos = reserva.pedidos.filter(pagado=False)
    total_pedidos = pedidos.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
    total_final = reserva.total_hospedaje + total_pedidos - reserva.adelanto

    if request.method == 'POST':
        # Crear la venta formal
        venta = Venta.objects.create(
            cliente=reserva.cliente,
            usuario=request.user,
            tipo_comprobante=request.POST.get('tipo_comprobante', 'BOLETA'),
            serie='F001' if request.POST.get('tipo_comprobante') == 'FACTURA' else 'B001',
            numero=Venta.objects.count() + 1,
            total=total_final,
            forma_pago=request.POST.get('forma_pago', 'CONTADO'),
            modulo='HOTEL'
        )
        
        # Detalle: Hospedaje
        producto_hospedaje, _ = Producto.objects.get_or_create(
            codigo='HOSP-01', 
            defaults={'nombre': 'Servicio de Hospedaje', 'precio_venta': reserva.total_hospedaje, 'categoria': 'HOSPEDAJE'}
        )
        DetalleVenta.objects.create(
            venta=venta,
            producto=producto_hospedaje,
            cantidad=1,
            precio_unitario=reserva.total_hospedaje,
            subtotal=reserva.total_hospedaje
        )
        
        # Detalle: Pedidos
        for p in pedidos:
            DetalleVenta.objects.create(
                venta=venta,
                producto=p.producto,
                cantidad=p.cantidad,
                precio_unitario=p.precio_unitario,
                subtotal=p.subtotal
            )
            p.pagado = True
            p.save()
        
        # Finalizar reserva
        reserva.estado = 'CHECKOUT'
        reserva.save()
        reserva.habitacion.estado = 'LIMPIEZA'
        reserva.habitacion.save()
        
        messages.success(request, "Checkout finalizado y venta generada.")
        return redirect('ventas_list')

    context = {
        'reserva': reserva,
        'pedidos': pedidos,
        'total_pedidos': total_pedidos,
        'total_final': total_final,
        'titulo': 'Checkout / Facturación de Estadía'
    }
    return render(request, 'gestion/checkout_form.html', context)

@login_required
def reserva_detalle(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    pedidos = reserva.pedidos.all().order_by('-fecha_pedido')
    total_pedidos = pedidos.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
    total_final = reserva.total_hospedaje + total_pedidos - reserva.adelanto
    
    context = {
        'reserva': reserva,
        'pedidos': pedidos,
        'total_pedidos': total_pedidos,
        'total_final': total_final,
        'titulo': f'Detalle de Estadía - Hab. {reserva.habitacion.numero}'
    }
    return render(request, 'gestion/reserva_detalle.html', context)

@login_required
def reserva_finalizar_temprano(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    if reserva.estado == 'CHECKIN':
        reserva.fecha_salida = timezone.now()
        # Recalcular total hospedaje basado en el nuevo tiempo
        diff = reserva.fecha_salida - reserva.fecha_ingreso
        noches = max(1, diff.days + (1 if diff.seconds > 3600 else 0)) # Al menos 1 hora cuenta como fracción
        reserva.total_hospedaje = noches * reserva.habitacion.precio_noche
        reserva.save()
        messages.info(request, f"Estadía finalizada antes de tiempo. Se han calculado {noches} noches.")
    return redirect('reserva_checkout', reserva_id=reserva.id)

@login_required
def reserva_delete(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    reserva.activo = False
    reserva.save()
    
    # Si estaba ocupada, liberar la habitación
    if reserva.estado == 'CHECKIN':
        reserva.habitacion.estado = 'DISPONIBLE'
        reserva.habitacion.save()
        
    messages.success(request, "Registro de recepción movido al historial.")
    return redirect('reserva_list')

@login_required
def reserva_restore(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    reserva.activo = True
    reserva.save()
    
    # Si vuelve a estar activa y no ha hecho checkout, volver a ocupar la habitación
    if reserva.estado == 'CHECKIN':
        reserva.habitacion.estado = 'OCUPADA'
        reserva.habitacion.save()
        
    messages.success(request, "Registro restaurado correctamente.")
    return redirect('reserva_list')

@login_required
def reserva_hacer_checkin(request, pk):
    """Convierte una reserva PENDIENTE en CHECKIN efectivo (huésped llega)."""
    reserva = get_object_or_404(Reserva, pk=pk, activo=True)
    if reserva.estado == 'PENDIENTE':
        reserva.estado = 'CHECKIN'
        reserva.save()
        reserva.habitacion.estado = 'OCUPADA'
        reserva.habitacion.save()
        messages.success(request, f"Check-in realizado para {reserva.cliente.nombre_razon_social} – Hab. {reserva.habitacion.numero}.")
    else:
        messages.warning(request, "Esta reserva no está en estado Pendiente.")
    return redirect('reserva_list')

@login_required
def mesa_abrir(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    if mesa.estado in ['DISPONIBLE', 'RESERVADA']:
        cliente, _ = Cliente.objects.get_or_create(
            numero_documento='00000000', 
            defaults={'nombre_razon_social': 'Público General', 'tipo_documento': 'DNI', 'modulo': 'RESTAURANTE'}
        )
        
        venta = Venta.objects.create(
            cliente=cliente,
            usuario=request.user,
            tipo_comprobante='TICKET',
            serie='COM',
            numero=Venta.objects.filter(serie='COM').count() + 1,
            total=0,
            es_comanda=True,
            mesa=mesa,
            modulo='RESTAURANTE'
        )
        mesa.estado = 'OCUPADA'
        mesa.save()
        messages.success(request, f"Mesa {mesa.numero} abierta.")
    return redirect('mesa_detalle', mesa_id=mesa.id)

@login_required
def mesa_detalle(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    comanda = Venta.objects.filter(mesa=mesa, es_comanda=True).last()
    productos = Producto.objects.filter(categoria='RESTAURANTE', activo=True)
    
    sillas_usadas = []
    if comanda:
        sillas_usadas = comanda.detalles.values_list('silla', flat=True).distinct().order_by('silla')
        # Filter out None values from sillas_usadas
        sillas_usadas = [s for s in sillas_usadas if s is not None]
        
    return render(request, 'gestion/mesa_detalle.html', {
        'mesa': mesa,
        'comanda': comanda,
        'productos': productos,
        'sillas_usadas': sillas_usadas,
        'sillas_range': range(1, mesa.capacidad + 1),
        'titulo': f'Mesa {mesa.numero}'
    })


@login_required
def mesa_agregar_item(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        cantidad = int(request.POST.get('cantidad', 1))
        silla = request.POST.get('silla') or None
        producto = get_object_or_404(Producto, id=producto_id)
        observaciones = request.POST.get('observaciones', '').strip() or None
        incluido_tour = request.POST.get('incluido_tour') in ['on', 'true']

        # Cada combinación producto+silla+observaciones+incluido_tour es un ítem único
        if silla:
            detalle, created = DetalleVenta.objects.get_or_create(
                venta=venta,
                producto=producto,
                silla=silla,
                observaciones=observaciones,
                incluido_tour=incluido_tour,
                defaults={
                    'cantidad': 0,
                    'precio_unitario': producto.precio_venta,
                    'subtotal': 0,
                    'mozo': request.user,
                }
            )
        else:
            detalle, created = DetalleVenta.objects.get_or_create(
                venta=venta,
                producto=producto,
                silla=None,
                observaciones=observaciones,
                incluido_tour=incluido_tour,
                defaults={
                    'cantidad': 0,
                    'precio_unitario': producto.precio_venta,
                    'subtotal': 0,
                    'mozo': request.user,
                }
            )

        detalle.cantidad += cantidad
        if incluido_tour:
            detalle.subtotal = Decimal('0.00')
        else:
            detalle.subtotal = detalle.cantidad * detalle.precio_unitario
            
        if not detalle.mozo:
            detalle.mozo = request.user
        detalle.save()

        # Crear EstadoComanda si no existe
        EstadoComanda.objects.get_or_create(detalle=detalle)

        venta.total = sum(d.subtotal for d in venta.detalles.all())
        venta.save()

    return redirect('mesa_detalle', mesa_id=venta.mesa.id)

@login_required
def mesa_cerrar(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    if request.method == 'POST':
        tipo = request.POST.get('tipo_comprobante', 'BOLETA')
        pago_tarjeta_checked = request.POST.get('pago_tarjeta') in ['on', 'true']
        silla_cobro = request.POST.get('silla_cobro') or None
        
        if silla_cobro:
            # Cobrar solo los ítems de esa silla
            silla_num = int(silla_cobro)
            detalles_silla = venta.detalles.filter(silla=silla_num)
            
            if not detalles_silla.exists():
                messages.warning(request, f"No hay pedidos para la silla {silla_num}.")
                return redirect('mesa_detalle', mesa_id=venta.mesa.id)
                
            # Crear una nueva Venta para esta silla
            nueva_venta = Venta.objects.create(
                cliente=venta.cliente,
                usuario=request.user,
                tipo_comprobante=tipo,
                serie='COM_S' if tipo == 'TICKET' else ('B001' if tipo == 'BOLETA' else 'F001'),
                numero=Venta.objects.filter(tipo_comprobante=tipo).count() + 1,
                total=0,
                forma_pago='CONTADO',
                es_comanda=False,
                pago_tarjeta=pago_tarjeta_checked,
                modulo='RESTAURANTE',
                mesa=venta.mesa,
                encargado_guia=venta.encargado_guia
            )
            
            # Mover detalles a la nueva venta
            for d in detalles_silla:
                d.venta = nueva_venta
                d.save()
                
            # Calcular totales de la nueva venta
            total_productos = sum(d.subtotal for d in nueva_venta.detalles.all())
            if pago_tarjeta_checked:
                recargo = total_productos * Decimal('0.05')
                nueva_venta.recargo_tarjeta = recargo
                nueva_venta.total = total_productos + recargo
            else:
                nueva_venta.recargo_tarjeta = Decimal('0.00')
                nueva_venta.total = total_productos
            nueva_venta.save()
            
            # Recalcular el total de la comanda original
            venta.total = sum(d.subtotal for d in venta.detalles.all())
            venta.save()
            
            # Si ya no quedan detalles en la comanda original, cerrarla y liberar mesa
            if not venta.detalles.exists():
                venta.es_comanda = False
                venta.save()
                if venta.mesa:
                    venta.mesa.estado = 'DISPONIBLE'
                    venta.mesa.save()
                messages.success(request, f"Cuenta de Silla {silla_num} cerrada. Mesa liberada por completo.")
                return redirect('restaurante')
            else:
                messages.success(request, f"Cuenta de Silla {silla_num} cerrada con éxito. La mesa continúa abierta.")
                return redirect('mesa_detalle', mesa_id=venta.mesa.id)
        else:
            # Proceso original (Toda la Mesa)
            venta.es_comanda = False
            venta.tipo_comprobante = tipo
            venta.pago_tarjeta = pago_tarjeta_checked
            
            total_productos = sum(d.subtotal for d in venta.detalles.all())
            if pago_tarjeta_checked:
                recargo = total_productos * Decimal('0.05')
                venta.recargo_tarjeta = recargo
                venta.total = total_productos + recargo
            else:
                venta.recargo_tarjeta = Decimal('0.00')
                venta.total = total_productos
                
            venta.save()
            
            if venta.mesa:
                venta.mesa.estado = 'DISPONIBLE'
                venta.mesa.save()
                
            messages.success(request, f"Cuenta de Mesa {venta.mesa.numero} cerrada.")
            return redirect('restaurante')
            
    return redirect('restaurante')


@login_required
def mesas_list(request):
    mesas = Mesa.objects.all().order_by('numero')
    return render(request, 'gestion/mesas_list.html', {'mesas': mesas})

@login_required
def mesa_create(request):
    if request.method == 'POST':
        form = MesaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Mesa creada correctamente.")
            return redirect('mesas_list')
    else:
        form = MesaForm()
    return render(request, 'gestion/mesa_form.html', {'form': form, 'titulo': 'Nueva Mesa'})

@login_required
def mesa_update(request, pk):
    mesa = get_object_or_404(Mesa, pk=pk)
    if request.method == 'POST':
        form = MesaForm(request.POST, instance=mesa)
        if form.is_valid():
            form.save()
            messages.success(request, "Mesa actualizada correctamente.")
            return redirect('mesas_list')
    else:
        form = MesaForm(instance=mesa)
    return render(request, 'gestion/mesa_form.html', {'form': form, 'titulo': 'Editar Mesa'})

@login_required
def mesa_delete(request, pk):
    mesa = get_object_or_404(Mesa, pk=pk)
    mesa.delete()
    messages.success(request, "Mesa eliminada correctamente.")
    return redirect('mesas_list')

@login_required
def carta_lista(request):
    platos = Producto.objects.filter(categoria='RESTAURANTE').order_by('subcategoria', 'nombre')
    categorias = {}
    for plato in platos:
        cat = plato.subcategoria or 'Otros'
        if cat not in categorias:
            categorias[cat] = []
        categorias[cat].append(plato)
        
    return render(request, 'gestion/carta_lista.html', {
        'categorias': categorias,
        'titulo': 'Carta / Menú'
    })

@login_required
def ticket_print(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    config = ConfiguracionEmpresa.objects.first()
    silla = request.GET.get('silla')  # Filtro opcional por silla
    detalles = venta.detalles.all()
    if silla:
        detalles = detalles.filter(silla=silla)
    
    sillas_usadas = venta.detalles.values_list('silla', flat=True).distinct().order_by('silla')
    
    # Calcular subtotal de los detalles (excluyendo incluidos en tour)
    subtotal_silla = sum(d.subtotal for d in detalles)
    
    # Calcular comisión total del guía
    total_comisiones = sum(d.cantidad * (d.producto.comision or Decimal('0.00')) for d in detalles)
    
    # Recargo de tarjeta del 5%
    pago_tarjeta = venta.pago_tarjeta
    if pago_tarjeta:
        recargo_tarjeta_silla = subtotal_silla * Decimal('0.05')
        total_silla_final = subtotal_silla + recargo_tarjeta_silla
    else:
        recargo_tarjeta_silla = Decimal('0.00')
        total_silla_final = subtotal_silla
        
    return render(request, 'gestion/ticket_print.html', {
        'venta': venta,
        'config': config,
        'detalles_filtrados': detalles,
        'silla_filtro': silla,
        'sillas_usadas': sillas_usadas,
        'subtotal_silla': subtotal_silla,
        'recargo_tarjeta_silla': recargo_tarjeta_silla,
        'total_silla_final': total_silla_final,
        'pago_tarjeta': pago_tarjeta,
        'total_comisiones': total_comisiones,
    })



@login_required
def mesa_guardar_guia(request, venta_id):
    """Guarda el nombre del guía/encargado de una comanda."""
    if request.method == 'POST':
        venta = get_object_or_404(Venta, id=venta_id)
        venta.encargado_guia = request.POST.get('encargado_guia', '').strip() or None
        venta.save()
        messages.success(request, "Guía/encargado guardado.")
    return redirect('mesa_detalle', mesa_id=venta.mesa.id)


@login_required
def jefe_mozos_panel(request):
    """Panel del Jefe de Mozos: ve todas las mesas, comandas activas y puede gestionar pedidos."""
    mesas = Mesa.objects.all().order_by('numero')
    comandas_activas = Venta.objects.filter(es_comanda=True).select_related('mesa', 'usuario')
    
    # Estadísticas rápidas
    total_mesas = mesas.count()
    mesas_ocupadas = mesas.filter(estado='OCUPADA').count()
    total_items_pendientes = EstadoComanda.objects.filter(
        estado='PENDIENTE',
        detalle__venta__es_comanda=True
    ).count()
    
    # Detalles pendientes agrupados por mesa
    detalles_pendientes = DetalleVenta.objects.filter(
        venta__es_comanda=True,
        estado_comanda__estado='PENDIENTE'
    ).select_related('venta', 'venta__mesa', 'producto', 'mozo', 'estado_comanda').order_by('venta__mesa__numero', 'id')
    
    return render(request, 'gestion/jefe_mozos_panel.html', {
        'mesas': mesas,
        'comandas_activas': comandas_activas,
        'total_mesas': total_mesas,
        'mesas_ocupadas': mesas_ocupadas,
        'total_items_pendientes': total_items_pendientes,
        'detalles_pendientes': detalles_pendientes,
        'titulo': 'Panel Jefe de Mozos',
    })

# ============================================================
# VISTAS DE ROLES ESPECIALES
# ============================================================

SUBCATEGORIAS_BEBIDAS = ['BEBIDA', 'LICOR']
SUBCATEGORIAS_COCINA = ['ENTRADA', 'PLATO_FONDO', 'POSTRE', 'OTRO']

@login_required
def cocina_view(request):
    """Panel exclusivo para el usuario de Cocina. Solo ve platos (no bebidas)."""
    # Pedidos activos: comandas abiertas con ítems de cocina (todo lo de restaurante excepto bebidas)
    detalles_cocina = DetalleVenta.objects.filter(
        venta__es_comanda=True,
        producto__categoria='RESTAURANTE',
    ).exclude(
        producto__subcategoria__in=SUBCATEGORIAS_BEBIDAS
    ).exclude(
        estado_comanda__estado='LISTO'
    ).select_related('venta', 'venta__mesa', 'producto', 'mozo', 'estado_comanda').order_by('venta__mesa__numero', 'id')

    # Agrupar por mesa/comanda
    pedidos_por_mesa = {}
    for detalle in detalles_cocina:
        mesa_key = f"Mesa {detalle.venta.mesa.numero}" if detalle.venta.mesa else "Sin Mesa"
        if mesa_key not in pedidos_por_mesa:
            pedidos_por_mesa[mesa_key] = {
                'mesa': detalle.venta.mesa,
                'venta_id': detalle.venta.id,
                'items': []
            }
        try:
            estado = detalle.estado_comanda.estado
        except:
            estado = 'PENDIENTE'
            EstadoComanda.objects.get_or_create(detalle=detalle)
        pedidos_por_mesa[mesa_key]['items'].append({
            'detalle': detalle,
            'estado': estado,
            'mozo_nombre': detalle.mozo.get_full_name() or detalle.mozo.username if detalle.mozo else 'Mozo',
        })

    return render(request, 'gestion/cocina_panel.html', {
        'pedidos_por_mesa': pedidos_por_mesa,
        'titulo': 'Panel de Cocina',
    })


@login_required
def barra_view(request):
    """Panel exclusivo para el usuario de Barra. Solo ve bebidas y licores."""
    detalles_barra = DetalleVenta.objects.filter(
        venta__es_comanda=True,
        producto__subcategoria__in=SUBCATEGORIAS_BEBIDAS,
        producto__categoria='RESTAURANTE',
    ).exclude(
        estado_comanda__estado='LISTO'
    ).select_related('venta', 'venta__mesa', 'producto', 'mozo', 'estado_comanda').order_by('venta__mesa__numero', 'id')

    pedidos_por_mesa = {}
    for detalle in detalles_barra:
        mesa_key = f"Mesa {detalle.venta.mesa.numero}" if detalle.venta.mesa else "Sin Mesa"
        if mesa_key not in pedidos_por_mesa:
            pedidos_por_mesa[mesa_key] = {
                'mesa': detalle.venta.mesa,
                'venta_id': detalle.venta.id,
                'items': []
            }
        try:
            estado = detalle.estado_comanda.estado
        except:
            estado = 'PENDIENTE'
            EstadoComanda.objects.get_or_create(detalle=detalle)
        pedidos_por_mesa[mesa_key]['items'].append({
            'detalle': detalle,
            'estado': estado,
            'mozo_nombre': detalle.mozo.get_full_name() or detalle.mozo.username if detalle.mozo else 'Mozo',
        })

    return render(request, 'gestion/barra_panel.html', {
        'pedidos_por_mesa': pedidos_por_mesa,
        'titulo': 'Panel de Barra',
    })


@login_required
def actualizar_estado_comanda(request, detalle_id):
    """Actualiza el estado de un ítem de comanda (cocina o barra). Vía POST."""
    if request.method == 'POST':
        detalle = get_object_or_404(DetalleVenta, id=detalle_id)
        nuevo_estado = request.POST.get('estado')
        estados_validos = ['PENDIENTE', 'ACEPTADO', 'EN_PREPARACION', 'LISTO']
        if nuevo_estado in estados_validos:
            comanda, _ = EstadoComanda.objects.get_or_create(detalle=detalle)
            comanda.estado = nuevo_estado
            comanda.actualizado_por = request.user
            comanda.save()
        # Devolver a la vista correcta según el rol
        try:
            rol = request.user.perfil.rol
        except:
            rol = 'ADMIN'
        if rol == 'BARRA':
            return redirect('barra')
        return redirect('cocina')
    return redirect('cocina')


@login_required
def limpieza_view(request):
    """Panel exclusivo para el usuario de Limpieza. Ve habitaciones en estado LIMPIEZA."""
    try:
        rol = request.user.perfil.rol
    except:
        rol = 'ADMIN'
        
    habitaciones_limpieza = Habitacion.objects.filter(
        estado='LIMPIEZA', activo=True
    ).order_by('numero')

    if rol == 'LIMPIEZA':
        habitaciones_limpieza = habitaciones_limpieza.filter(personal_limpieza=request.user)

    return render(request, 'gestion/limpieza_panel.html', {
        'habitaciones': habitaciones_limpieza,
        'titulo': 'Panel de Limpieza',
    })


@login_required
def actualizar_estado_limpieza(request, hab_id):
    """Cambia el estado de una habitación desde el panel de limpieza."""
    if request.method == 'POST':
        habitacion = get_object_or_404(Habitacion, id=hab_id)
        nuevo_subestado = request.POST.get('subestado')
        
        if nuevo_subestado in ['RECIBIDO', 'LIMPIANDO']:
            habitacion.subestado_limpieza = nuevo_subestado
            habitacion.save()
            messages.info(request, f"Habitación #{habitacion.numero} marcada en estado: {habitacion.get_subestado_limpieza_display()}.")
        elif nuevo_subestado == 'TERMINADO':
            habitacion.estado = 'DISPONIBLE'
            habitacion.subestado_limpieza = 'PENDIENTE'
            habitacion.personal_limpieza = None
            habitacion.save()
            messages.success(request, f"Habitación #{habitacion.numero} finalizada y marcada como Disponible.")
    return redirect('limpieza')


@login_required
def habitacion_liberar(request, hab_id):
    """El admin libera una habitación para limpieza."""
    habitacion = get_object_or_404(Habitacion, id=hab_id)
    habitacion.estado = 'LIMPIEZA'
    if request.method == 'POST':
        personal_id = request.POST.get('personal_limpieza')
        if personal_id:
            from django.contrib.auth.models import User
            personal = get_object_or_404(User, id=personal_id)
            habitacion.personal_limpieza = personal
        else:
            habitacion.personal_limpieza = None
    habitacion.subestado_limpieza = 'PENDIENTE'
    habitacion.save()
    messages.success(request, f"Habitación #{habitacion.numero} asignada/liberada para limpieza.")
    return redirect('habitaciones_list')


@login_required
def habitacion_detalle_cliente(request, reserva_id):
    """Retorna los datos del cliente de una reserva activa (para modal)."""
    reserva = get_object_or_404(Reserva, id=reserva_id)
    pedidos = reserva.pedidos.all().order_by('-fecha_pedido')
    total_pedidos = pedidos.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
    total_final = reserva.total_hospedaje + total_pedidos - reserva.adelanto
    return render(request, 'gestion/habitacion_cliente_modal.html', {
        'reserva': reserva,
        'pedidos': pedidos,
        'total_pedidos': total_pedidos,
        'total_final': total_final,
    })

